import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import comm


def append_csv_to_existing_sheet(csv_file_path, excel_template_path, sheet_name):
    # Read the CSV file into a DataFrame
    df = pd.read_csv(csv_file_path)

    # Load the existing Excel workbook
    wb = load_workbook(excel_template_path)

    # Check if the sheet exists
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")

    # Find the next available row in the sheet
    next_row = ws.max_row + 1

    # Append the DataFrame to the Excel sheet starting from the next available row
    for row in dataframe_to_rows(df, index=False, header=ws.max_row == 0):
        for col_idx, cell_value in enumerate(row, 1):
            ws.cell(row=next_row, column=col_idx, value=cell_value)
        next_row += 1

    # Save the changes to the workbook
    wb.save(excel_template_path)
    print(f"CSV data from {csv_file_path} has been appended to sheet '{sheet_name}' in {excel_template_path}")


# Example usage:
#sheet_name ='Clients'
#sheet_name="Contacts"
#sheet_name = 'ClientContacts'
#sheet_name = 'ClientGroups'
#sheet_name = 'Groups'
sheet_name = 'Services'
csv_file_path = os.path.join(comm.folder_4_mysql_results,sheet_name+'.csv')  # Path to your CSV file
excel_template_path = comm.template_file  # Path to your Excel template

append_csv_to_existing_sheet(csv_file_path, excel_template_path, sheet_name)
